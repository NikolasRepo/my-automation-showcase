from datetime import date
from io import BytesIO
from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select
from sqlalchemy.orm import selectinload
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from app.core.database import get_db
from app.core.security import require_role
from app.models.user import User, UserRole
from app.models.inspection import Inspection, InspectionResult

router = APIRouter(prefix="/reports", tags=["reports"])


@router.get("/export")
async def export_excel(
    date_from: date = Query(...),
    date_to: date = Query(...),
    db: AsyncSession = Depends(get_db),
    _: User = Depends(require_role(UserRole.leader, UserRole.admin)),
):
    result = await db.execute(
        select(Inspection)
        .where(
            Inspection.inspection_date >= date_from,
            Inspection.inspection_date <= date_to,
        )
        .options(
            selectinload(Inspection.operator),
            selectinload(Inspection.production_line),
            selectinload(Inspection.results).selectinload(InspectionResult.checklist_item),
            selectinload(Inspection.results).selectinload(InspectionResult.station),
        )
        .order_by(Inspection.inspection_date, Inspection.submitted_at)
    )
    inspections = result.scalars().all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Inspection Report"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E79")
    flag_fill = PatternFill("solid", fgColor="FFE0E0")
    center = Alignment(horizontal="center")

    headers = [
        "Date", "Shift", "Production Line", "Station", "Operator",
        "Status", "Submitted At", "Check Item", "Category",
        "Type", "Pass/Fail", "Value", "Unit", "Flagged", "Notes"
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    row = 2
    for insp in inspections:
        for res in insp.results:
            item = res.checklist_item
            ws.cell(row=row, column=1, value=str(insp.inspection_date))
            ws.cell(row=row, column=2, value=insp.shift.value)
            ws.cell(row=row, column=3, value=insp.production_line.name if insp.production_line else "")
            ws.cell(row=row, column=4, value=res.station.name if res.station else "")
            ws.cell(row=row, column=5, value=insp.operator.full_name if insp.operator else "")
            ws.cell(row=row, column=6, value=insp.status.value)
            ws.cell(row=row, column=7, value=str(insp.submitted_at) if insp.submitted_at else "")
            ws.cell(row=row, column=8, value=item.name)
            ws.cell(row=row, column=9, value=item.category or "")
            ws.cell(row=row, column=10, value=item.data_type.value)
            ws.cell(row=row, column=11, value="Pass" if res.pass_fail else "Fail" if res.pass_fail is False else "")
            ws.cell(row=row, column=12, value=res.numeric_value)
            ws.cell(row=row, column=13, value=item.unit or "")
            ws.cell(row=row, column=14, value="YES" if res.flagged else "")
            ws.cell(row=row, column=15, value=res.notes or "")
            if res.flagged:
                for col in range(1, 16):
                    ws.cell(row=row, column=col).fill = flag_fill
            row += 1

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 18

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)

    filename = f"inspections_{date_from}_to_{date_to}.xlsx"
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )
