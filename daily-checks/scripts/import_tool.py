"""
Daily Checks Portal — Excel Import Tool
========================================
Reads a checks_summary.xlsx file, previews all changes, and imports
them into the Daily Checks Portal via the API.

Usage:
    python import_tool.py checks_summary.xlsx
    python import_tool.py checks_summary.xlsx --url http://192.168.1.50
    python import_tool.py checks_summary.xlsx --url http://192.168.1.50 --username admin --password admin

Requirements:
    pip install requests openpyxl

The script will:
    1. Connect to the Daily Checks Portal
    2. Read your Excel file
    3. Compare against what already exists in the system
    4. Show a full preview of all changes and warnings
    5. Ask for confirmation before making any changes
"""

import argparse
import getpass
import sys
import os

try:
    import requests
except ImportError:
    print("ERROR: 'requests' library not installed.")
    print("Run: pip install requests")
    sys.exit(1)

try:
    import openpyxl
except ImportError:
    print("ERROR: 'openpyxl' library not installed.")
    print("Run: pip install openpyxl")
    sys.exit(1)


# ─── Configuration ────────────────────────────────────────────────────────────

DEFAULT_URL = "http://localhost"


# ─── Data type parsing ────────────────────────────────────────────────────────

def parse_data_type(value):
    if not value:
        return "pass_fail", None
    v = str(value).strip().lower().replace(" ", "_").replace("/", "_")
    if v in ("pass_fail", "pass", "fail", "pf"):
        return "pass_fail", None
    elif v in ("ok_ng", "ok", "ng", "okng"):
        return "ok_ng", None
    elif v in ("numeric", "num", "measurement", "number", "value"):
        return "numeric", None
    else:
        return "pass_fail", f"Unknown check type '{value}' — defaulting to pass_fail"


# ─── API client ───────────────────────────────────────────────────────────────

class PortalClient:
    def __init__(self, base_url):
        self.base_url = base_url.rstrip("/")
        self.token = None
        self.session = requests.Session()

    def login(self, username, password):
        url = f"{self.base_url}/api/auth/token"
        try:
            resp = self.session.post(url, data={"username": username, "password": password}, timeout=10)
        except requests.exceptions.ConnectionError:
            print(f"\nERROR: Cannot connect to {self.base_url}")
            print("Make sure the Daily Checks Portal is running and the URL is correct.")
            sys.exit(1)
        except requests.exceptions.Timeout:
            print(f"\nERROR: Connection to {self.base_url} timed out.")
            sys.exit(1)

        if resp.status_code == 401:
            print("\nERROR: Incorrect username or password.")
            sys.exit(1)
        if resp.status_code != 200:
            print(f"\nERROR: Login failed ({resp.status_code}): {resp.text}")
            sys.exit(1)

        self.token = resp.json()["access_token"]
        self.session.headers.update({"Authorization": f"Bearer {self.token}"})
        return resp.json()["user"]

    def get(self, path, params=None):
        resp = self.session.get(f"{self.base_url}/api{path}", params=params, timeout=10)
        resp.raise_for_status()
        return resp.json()

    def post(self, path, data):
        resp = self.session.post(f"{self.base_url}/api{path}", json=data, timeout=10)
        resp.raise_for_status()
        return resp.json()


# ─── Excel reading ────────────────────────────────────────────────────────────

def read_excel(filepath):
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active

    headers = [str(cell.value).strip().lower() if cell.value else "" for cell in ws[1]]

    col = {}
    for i, h in enumerate(headers):
        if "line" in h and "sub" not in h:
            col["line"] = i
        elif ("process" in h or "station" in h) and "sub" not in h:
            col["station"] = i
        elif "sub" in h:
            col["sub_station"] = i
        elif "type" in h and "check" not in h:
            pass  # type (e/q) — ignored
        elif "description" in h or ("check" in h and "d" in h):
            col["name"] = i
        elif "check type" in h or ("check" in h and "type" in h):
            col["data_type"] = i

    required = ["line", "station", "name", "data_type"]
    missing = [r for r in required if r not in col]
    if missing:
        print(f"\nERROR: Missing required columns in Excel file: {missing}")
        print(f"Found columns: {headers}")
        sys.exit(1)

    rows = []
    warnings = []
    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        line = str(row[col["line"]]).strip() if row[col["line"]] else None
        station = str(row[col["station"]]).strip() if row[col["station"]] else None
        name = str(row[col["name"]]).strip() if row[col["name"]] else None

        if not line or not station:
            continue

        if not name:
            warnings.append(f"Row {row_num}: Blank check description — row will be skipped")
            continue

        data_type_raw = row[col["data_type"]] if col.get("data_type") else None
        data_type, warn = parse_data_type(data_type_raw)
        if warn:
            warnings.append(f"Row {row_num}: {warn}")

        sub_station = None
        if col.get("sub_station") and row[col["sub_station"]]:
            sub_station = str(row[col["sub_station"]]).strip()

        rows.append({
            "line": line,
            "station": station,
            "sub_station": sub_station,
            "name": name,
            "data_type": data_type,
        })

    return rows, warnings


# ─── Preview builder ──────────────────────────────────────────────────────────

def build_preview(rows, existing_lines):
    # Build lookup of existing data
    existing_line_map = {l["name"]: l for l in existing_lines}
    existing_station_map = {}
    existing_item_map = {}

    for line in existing_lines:
        for station in line.get("stations", []):
            key = f"{line['name']}|{station['name']}"
            existing_station_map[key] = station
            for item in station.get("checklist_items", []):
                item_key = f"{line['name']}|{station['name']}|{item['name']}"
                existing_item_map[item_key] = item

    # Build structured changes
    changes = {}  # line_name -> { stations: { station_name -> [items] } }
    for row in rows:
        line_name = row["line"]
        station_name = row["station"]
        item_name = row["name"]

        if line_name not in changes:
            changes[line_name] = {"stations": {}}
        if station_name not in changes[line_name]["stations"]:
            changes[line_name]["stations"][station_name] = []
        changes[line_name]["stations"][station_name].append(row)

    # Determine what is new vs existing
    preview = []
    total_new_lines = 0
    total_new_stations = 0
    total_new_items = 0
    total_skipped = 0

    for line_name, line_data in changes.items():
        line_is_new = line_name not in existing_line_map
        if line_is_new:
            total_new_lines += 1

        line_preview = {
            "name": line_name,
            "is_new": line_is_new,
            "stations": []
        }

        for station_name, items in line_data["stations"].items():
            station_key = f"{line_name}|{station_name}"
            station_is_new = station_key not in existing_station_map

            if station_is_new:
                total_new_stations += 1

            new_items = []
            skipped_items = []
            for item in items:
                item_key = f"{line_name}|{station_name}|{item['name']}"
                if item_key in existing_item_map:
                    skipped_items.append(item)
                    total_skipped += 1
                else:
                    new_items.append(item)
                    total_new_items += 1

            line_preview["stations"].append({
                "name": station_name,
                "is_new": station_is_new,
                "new_items": new_items,
                "skipped_items": skipped_items,
            })

        preview.append(line_preview)

    return preview, total_new_lines, total_new_stations, total_new_items, total_skipped


# ─── Display preview ──────────────────────────────────────────────────────────

def display_preview(preview, warnings, total_new_lines, total_new_stations, total_new_items, total_skipped):
    print("\n" + "═" * 50)
    print("  Daily Checks Portal — Import Preview")
    print("═" * 50)

    if not any(
        item
        for line in preview
        for station in line["stations"]
        for item in station["new_items"]
    ) and not any(line["is_new"] for line in preview) and not any(
        station["is_new"] for line in preview for station in line["stations"]
    ):
        print("\n  Nothing new to import — all items already exist in the system.")
        if warnings:
            print("\nWARNINGS:")
            for w in warnings:
                print(f"  ⚠  {w}")
        print()
        return False

    print("\nCHANGES TO BE MADE:")
    print("─" * 50)

    for line in preview:
        line_label = "NEW LINE" if line["is_new"] else "EXISTING LINE"
        print(f"\n  {line_label}: \"{line['name']}\"")

        for station in line["stations"]:
            station_label = "NEW STATION" if station["is_new"] else "EXISTING STATION"
            new_count = len(station["new_items"])
            skip_count = len(station["skipped_items"])

            if new_count == 0 and skip_count > 0:
                print(f"    {station_label}: \"{station['name']}\" — no new items ({skip_count} already exist)")
                continue

            if new_count == 0:
                continue

            suffix = ""
            if skip_count > 0:
                suffix = f" ({skip_count} already exist, skipping)"

            print(f"    {station_label}: \"{station['name']}\" — {new_count} new item(s){suffix}")

            for item in station["new_items"]:
                type_label = {"pass_fail": "Pass/Fail", "ok_ng": "OK/NG", "numeric": "Numeric"}.get(item["data_type"], item["data_type"])
                sub = f" [{item['sub_station']}]" if item.get("sub_station") else ""
                print(f"      + {item['name']}{sub}  ({type_label})")

    if warnings:
        print("\nWARNINGS:")
        print("─" * 50)
        for w in warnings:
            print(f"  ⚠  {w}")

    print("\n" + "─" * 50)
    print(f"  {total_new_lines} new line(s)  |  {total_new_stations} new station(s)  |  {total_new_items} new item(s)")
    if total_skipped > 0:
        print(f"  {total_skipped} item(s) already exist and will be skipped")
    print()

    return total_new_items > 0 or total_new_lines > 0 or total_new_stations > 0


# ─── Import execution ─────────────────────────────────────────────────────────

def run_import(client, preview):
    print("Importing...")
    lines_created = 0
    stations_created = 0
    items_created = 0

    # Fetch current lines to get IDs
    existing_lines = client.get("/admin/lines")
    line_map = {l["name"]: l for l in existing_lines}

    for line_preview in preview:
        line_name = line_preview["name"]

        # Create line if new
        if line_preview["is_new"]:
            line = client.post("/admin/lines", {"name": line_name})
            line_map[line_name] = line
            lines_created += 1
            print(f"  Created line: {line_name}")

        line_id = line_map[line_name]["id"]

        # Build station map for this line
        station_map = {s["name"]: s for s in line_map[line_name].get("stations", [])}

        for station_preview in line_preview["stations"]:
            station_name = station_preview["name"]

            # Skip if no new items
            if not station_preview["new_items"]:
                continue

            # Create station if new
            if station_preview["is_new"]:
                existing = client.get("/admin/lines")
                line_data = next((l for l in existing if l["id"] == line_id), None)
                display_order = len(line_data.get("stations", [])) + 1 if line_data else 1
                station = client.post("/admin/stations", {
                    "production_line_id": line_id,
                    "name": station_name,
                    "display_order": display_order,
                })
                station_map[station_name] = station
                stations_created += 1
                print(f"    Created station: {station_name}")

            station_id = station_map[station_name]["id"]

            # Get current item count for display order
            existing_items = station_map[station_name].get("checklist_items", [])
            display_order = len(existing_items)

            # Create new items
            for item in station_preview["new_items"]:
                display_order += 1
                client.post("/admin/checklist-items", {
                    "station_id": station_id,
                    "name": item["name"],
                    "sub_station": item.get("sub_station"),
                    "data_type": item["data_type"],
                    "display_order": display_order,
                })
                items_created += 1

    print(f"\nImport complete:")
    print(f"  {lines_created} line(s) created")
    print(f"  {stations_created} station(s) created")
    print(f"  {items_created} checklist item(s) created")


# ─── Main ─────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Daily Checks Portal — Excel Import Tool")
    parser.add_argument("filepath", help="Path to your Excel file (e.g. checks_summary.xlsx)")
    parser.add_argument("--url", default=DEFAULT_URL, help=f"Portal URL (default: {DEFAULT_URL})")
    parser.add_argument("--username", default=None, help="Admin username")
    parser.add_argument("--password", default=None, help="Admin password")
    args = parser.parse_args()

    # Validate file exists
    if not os.path.exists(args.filepath):
        print(f"\nERROR: File not found: {args.filepath}")
        sys.exit(1)

    print(f"\nDaily Checks Portal Import Tool")
    print(f"Portal URL: {args.url}")
    print(f"File: {args.filepath}")

    # Get credentials if not provided
    username = args.username or input("\nUsername: ")
    password = args.password or getpass.getpass("Password: ")

    # Connect and login
    print("\nConnecting to portal...")
    client = PortalClient(args.url)
    user = client.login(username, password)
    print(f"Logged in as: {user['full_name']} ({user['role']})")

    # Check role
    if user["role"] not in ("admin", "leader"):
        print("\nERROR: You must be an admin or leader to import data.")
        sys.exit(1)

    # Fetch existing data
    print("Fetching existing data from portal...")
    existing_lines = client.get("/admin/lines")

    # Read Excel file
    print(f"Reading {args.filepath}...")
    rows, warnings = read_excel(args.filepath)
    print(f"Found {len(rows)} data rows")

    # Build preview
    preview, total_new_lines, total_new_stations, total_new_items, total_skipped = build_preview(rows, existing_lines)

    # Display preview
    has_changes = display_preview(preview, warnings, total_new_lines, total_new_stations, total_new_items, total_skipped)

    if not has_changes:
        sys.exit(0)

    # Confirmation prompt
    try:
        answer = input("Proceed with import? (yes/no): ").strip().lower()
    except KeyboardInterrupt:
        print("\n\nImport cancelled.")
        sys.exit(0)

    if answer not in ("yes", "y"):
        print("\nImport cancelled.")
        sys.exit(0)

    print()
    run_import(client, preview)
    print("\nDone. Refresh the Daily Checks Portal to see the new items.")


if __name__ == "__main__":
    main()
