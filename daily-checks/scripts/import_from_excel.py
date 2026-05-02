"""
Excel Import Script
Usage: python scripts/import_from_excel.py path/to/your/file.xlsx

This script reads your Excel file and imports production lines, stations,
and checklist items directly into the database.

Expected columns (row 1 headers):
  - file             : reference number (stored but not used in app)
  - line             : production line name
  - process / station: station name
  - sub-station      : sub-station name (optional)
  - type             : category e.g. "equipment check", "quality check"
  - check description: checklist item name
  - check type       : pass_fail, ok_ng, or numeric

Check type values accepted (case insensitive):
  pass_fail, pass/fail, pf        → DataType.pass_fail
  ok_ng, ok/ng, okng              → DataType.ok_ng
  numeric, num, measurement       → DataType.numeric
"""
import asyncio
import sys
import os
sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "backend"))

import openpyxl
from app.core.database import AsyncSessionLocal, engine, Base
from app.models.user import User
from app.models.inspection import ProductionLine, Station, ChecklistItem, DataType
from sqlalchemy import select
from sqlalchemy.orm import selectinload


def parse_data_type(value: str) -> DataType:
    """Convert Excel check type value to DataType enum."""
    if not value:
        return DataType.pass_fail
    v = str(value).strip().lower().replace(" ", "_").replace("/", "_")
    if v in ("pass_fail", "pass", "fail", "pf", "pass_fail"):
        return DataType.pass_fail
    elif v in ("ok_ng", "ok", "ng", "okng", "ok_ng"):
        return DataType.ok_ng
    elif v in ("numeric", "num", "measurement", "number", "value"):
        return DataType.numeric
    else:
        print(f"  WARNING: Unknown check type '{value}' — defaulting to pass_fail")
        return DataType.pass_fail


async def import_excel(filepath: str):
    print(f"Reading: {filepath}")

    wb = openpyxl.load_workbook(filepath)
    ws = wb.active

    # Read headers from row 1
    headers = [str(cell.value).strip().lower() if cell.value else "" for cell in ws[1]]
    print(f"Headers found: {headers}")

    # Map header names to column indices
    col = {}
    for i, h in enumerate(headers):
        if "line" in h and "sub" not in h:
            col["line"] = i
        elif "process" in h or "station" in h and "sub" not in h:
            col["station"] = i
        elif "sub" in h:
            col["sub_station"] = i
        elif "type" in h and "check" not in h:
            pass  # type (e/q) column — ignored, outdated
        elif "description" in h or "check d" in h:
            col["name"] = i
        elif "check type" in h or ("check" in h and "type" in h):
            col["data_type"] = i
        elif "file" in h:
            col["file"] = i

    print(f"Column mapping: {col}")

    # Validate required columns
    required = ["line", "station", "name", "data_type"]
    missing = [r for r in required if r not in col]
    if missing:
        print(f"ERROR: Missing required columns: {missing}")
        print("Please check your Excel headers match the expected format.")
        return

    # Read all data rows
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        line = str(row[col["line"]]).strip() if row[col["line"]] else None
        station = str(row[col["station"]]).strip() if row[col["station"]] else None
        name = str(row[col["name"]]).strip() if row[col["name"]] else None
        if not line or not station or not name:
            continue  # Skip empty rows
        rows.append({
            "line": line,
            "station": station,
            "sub_station": str(row[col["sub_station"]]).strip() if col.get("sub_station") and row[col["sub_station"]] else None,
            "name": name,
            "data_type": parse_data_type(row[col["data_type"]] if col.get("data_type") else None),
        })

    print(f"Found {len(rows)} data rows")

    async with engine.begin() as conn:
        await conn.run_sync(Base.metadata.create_all)

    async with AsyncSessionLocal() as db:
        lines_created = 0
        stations_created = 0
        items_created = 0

        # Cache to avoid duplicate lookups
        line_cache = {}
        station_cache = {}
        display_order_cache = {}

        for row in rows:
            line_name = row["line"]
            station_name = row["station"]

            # Get or create production line
            if line_name not in line_cache:
                result = await db.execute(
                    select(ProductionLine).where(ProductionLine.name == line_name)
                )
                line = result.scalar_one_or_none()
                if not line:
                    line = ProductionLine(name=line_name)
                    db.add(line)
                    await db.flush()
                    lines_created += 1
                    print(f"  Created line: {line_name}")
                line_cache[line_name] = line

            line = line_cache[line_name]

            # Get or create station
            station_key = f"{line_name}|{station_name}"
            if station_key not in station_cache:
                result = await db.execute(
                    select(Station).where(
                        Station.production_line_id == line.id,
                        Station.name == station_name,
                    )
                )
                station = result.scalar_one_or_none()
                if not station:
                    # Calculate display order for this station
                    station_order = len([k for k in station_cache if k.startswith(f"{line_name}|")]) + 1
                    station = Station(
                        production_line_id=line.id,
                        name=station_name,
                        display_order=station_order,
                    )
                    db.add(station)
                    await db.flush()
                    stations_created += 1
                    print(f"    Created station: {station_name}")
                station_cache[station_key] = station

            station = station_cache[station_key]

            # Calculate display order for this checklist item
            if station_key not in display_order_cache:
                display_order_cache[station_key] = 0
            display_order_cache[station_key] += 1

            # Create checklist item
            item = ChecklistItem(
                station_id=station.id,
                name=row["name"],
                sub_station=row["sub_station"],
                category=None,
                data_type=row["data_type"],
                display_order=display_order_cache[station_key],
            )
            db.add(item)
            items_created += 1

        await db.commit()

        print(f"\nImport complete:")
        print(f"  {lines_created} production line(s) created")
        print(f"  {stations_created} station(s) created")
        print(f"  {items_created} checklist item(s) created")
        print("\nNote: Run docker compose down -v and re-seed if you need a fresh start.")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python scripts/import_from_excel.py path/to/your/file.xlsx")
        sys.exit(1)
    filepath = sys.argv[1]
    if not os.path.exists(filepath):
        print(f"File not found: {filepath}")
        sys.exit(1)
    asyncio.run(import_excel(filepath))
